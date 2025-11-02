# %%
print("in this selenium code add in future")

# %%
#this is the code for csv file name to News.csv
import os

# Folder path where the CSV file is located
folder_path = 'Github/News'

# Loop through the files in the folder
for filename in os.listdir(folder_path):
    # Check if the file is a CSV file
    if filename.endswith('.csv'):
        # Full path to the old file
        old_file_path = os.path.join(folder_path, filename)
        
        # Full path to the new file name
        new_file_path = os.path.join(folder_path, 'News.csv')
        
        # Rename the file
        os.rename(old_file_path, new_file_path)
        print(f'Renamed "{filename}" to "News.csv"')
        break  # Stop after renaming the first CSV file found

# If no CSV file was found, you can add a message
else:
    print("No CSV file found in the specified folder.")


# %%
#new code for removing column FILE SIZE in csv file 22-05-2025
import pandas as pd

def remove_file_size_column(file_path: str) -> None:
    """
    Removes the 'FILE SIZE' column from the CSV file and overwrites the original file.

    Parameters
    ----------
    file_path : str
        Full path of the CSV file to modify.

    Returns
    -------
    None
        The original CSV file is updated in place.
    """
    # Read the CSV file
    df = pd.read_csv(file_path)

    # Check if 'FILE SIZE' column exists
    if "FILE SIZE" in df.columns:
        # Drop the column
        df = df.drop(columns=["FILE SIZE"])
        print("'FILE SIZE' column removed.")
    else:
        print("No 'FILE SIZE' column found. No changes made.")

    # Overwrite the original file with the updated data
    df.to_csv(file_path, index=False)
    print(f"File overwritten successfully: {file_path}")

# Example usage
file_path = 'Github/News/News.csv'
remove_file_size_column(file_path)


# %%
#2 remove zip and xml and excel 
import os
import pandas as pd

def remove_specific_rows(file_path):
    # Read the CSV file into a pandas DataFrame
    df = pd.read_csv(file_path)
    
    # Remove rows where the ATTACHMENT column ends with ".xml", ".zip", or ".xlsx"
    df = df[~df['ATTACHMENT'].str.endswith(('.xml', '.zip', '.xlsx'))]
    
    # Save the updated DataFrame to the same CSV file
    df.to_csv(file_path, index=False)

# Function to process all CSV files in a folder
def process_csv_files(folder_path):
    # Iterate over all files in the folder
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.csv'):
            file_path = os.path.join(folder_path, file_name)
            print("Processing:", file_path)
            
            # Read the CSV file
            df = pd.read_csv(file_path)
            
            # Remove rows where data in the "ATTACHMENT" column starts with "-"
            df = df[~df['ATTACHMENT'].str.startswith('-')]
            
            # Write the modified DataFrame back to the same CSV file
            df.to_csv(file_path, index=False)
            
            # Call the function to remove XML, ZIP, and XLSX rows
            remove_specific_rows(file_path)
            
            print("XML, ZIP, and XLSX rows removed successfully.")

# Folder path containing CSV files
folder_path = "Github/News"

# Call the function to process all CSV files in the folder
process_csv_files(folder_path)


# %%
#3 remove loss of certificate
import os
import pandas as pd

# Folder path
folder_path = "Github/News"

# List all files in the folder
file_names = os.listdir(folder_path)

# Iterate through each file in the folder
for file_name in file_names:
    # Check if the file is a CSV file
    if file_name.endswith('.csv'):
        # Construct the full file path
        file_path = os.path.join(folder_path, file_name)
        
        # Read the CSV file
        data = pd.read_csv(file_path)
        
        # Filter out rows containing "ESOP/ESOS/ESPS" or "Loss of Share Certificates" in the SUBJECT column
        filtered_data = data[~data['SUBJECT'].str.contains('ESOP/ESOS/ESPS|Loss of Share Certificates|Loss of share certificat|Loss of share certificate|ESOP/ESPS/SBEB Scheme|loss of share certificates')]
        
        # Overwrite the original file with the filtered data
        filtered_data.to_csv(file_path, index=False)
        
        print("Filtered data has been saved to", file_path)


# %%
#4 pdf links extract
import pandas as pd
import os
import glob

# Input folder path containing CSV files
input_folder_path =  "Github/News"

# Output folder path for saving modified CSV files
output_folder_path = "Github/News"

# Get a list of all CSV files in the input folder
csv_files = glob.glob(os.path.join(input_folder_path, '*.csv'))

# Iterate over each CSV file
for file_path in csv_files:
    # Read the CSV file
    df = pd.read_csv(file_path)
    
    # Extract filenames from the "ATTACHMENT" column
    df['LINKS'] = df['ATTACHMENT'].apply(lambda x: x.split('/')[-1])
    
    # Construct the output file path
    output_file_path = os.path.join(output_folder_path, os.path.basename(file_path))
    
    # Save the modified DataFrame to the output CSV file
    df.to_csv(output_file_path, index=False)
    
    # Print a message indicating the operation is done for each file
    print(f"Filenames extracted and saved to the 'links' column in the CSV file: {output_file_path}")


# %%
#5empty database create
import sqlite3
import os

# Define the folder path
folder_path = "Github/Database"

# Check if the folder exists, if not, create it
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Construct the full file path for the SQLite database
db_file_path = os.path.join(folder_path, 'News Analysis.db')

# Create an empty SQLite database file
with sqlite3.connect(db_file_path) as conn:
    print("Database file created successfully at:", db_file_path)

# Close the connection
conn.close()


# %%
#6 csv data add in database
import os
import pandas as pd
import sqlite3

# Folder path containing CSV files
csv_folder =  "Github/News"

# Database file path
db_file =  "Github/Database/News Analysis.db"

# Function to create database table from CSV file
def create_table_from_csv(csv_file, cursor):
    table_name = os.path.splitext(os.path.basename(csv_file))[0]  # Extract table name from CSV file name
    df = pd.read_csv(csv_file)  # Read CSV file into DataFrame
    try:
        df.columns = ['SYMBOL', 'COMPANY NAME', 'SUBJECT', 'DETAILS', 'BROADCAST DATE/TIME', 'RECEIPT', 'DISSEMINATION', 'DIFFERENCE', 'ATTACHMENT', 'LINKS']  # Rename columns
    except ValueError:
        pass  # Ignore length mismatch error
    # Add "ID" column as the first column with auto-generated numbers
    df.insert(0, 'ID', range(1, len(df) + 1))
    df.to_sql(table_name, conn, if_exists='replace', index=False)  # Write DataFrame to SQLite database as a table

# Connect to SQLite database
conn = sqlite3.connect(db_file)
cursor = conn.cursor()

# Iterate over CSV files in the folder
for file_name in os.listdir(csv_folder):
    if file_name.endswith('.csv'):
        csv_file_path = os.path.join(csv_folder, file_name)
        create_table_from_csv(csv_file_path, cursor)

# Commit changes and close connection
conn.commit()
conn.close()

print("Database successfully created from CSV files with 'ID' column.")



# %%
#7 filter the news
import sqlite3

# Connect to the first database
try:
    conn1 = sqlite3.connect("Github/Database/News Analysis.db")
    cursor1 = conn1.cursor()
except sqlite3.Error as e:
    print("Error connecting to the first database:", e)

# Connect to the second database
try:
    conn2 = sqlite3.connect("Github/My Stocks/MyStocks.db")
    cursor2 = conn2.cursor()
except sqlite3.Error as e:
    print("Error connecting to the second database:", e)

# Fetch symbols from the second database
try:
    cursor2.execute("SELECT SYMBOL FROM News")
    symbols = [row[0] for row in cursor2.fetchall()]
except sqlite3.Error as e:
    print("Error fetching symbols from the second database:", e)

# Delete rows from the first database where SYMBOL is not in symbols list
try:
    cursor1.execute("DELETE FROM News WHERE SYMBOL NOT IN ({})".format(','.join(['?']*len(symbols))), symbols)
    conn1.commit()
except sqlite3.Error as e:
    print("Error deleting rows from the first database:", e)
finally:
    conn1.close()
    conn2.close()

print("Operation completed successfully!")



# %%


# %%
#8this is the new code for csv data copy and paste in new excelfile 
import pandas as pd
import openpyxl

# Define the paths to your files
csv_file_path = 'Github/News/News.csv'
excel_folder_path = 'Github/Excel format news'
excel_file_name = 'Symbols.xlsx'
excel_file_path = f'{excel_folder_path}/{excel_file_name}'

# Load the CSV file
csv_data = pd.read_csv(csv_file_path)

# Extract the "ATTACHMENT" column data
attachment_data = csv_data['ATTACHMENT']

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

# Write the "ATTACHMENT" column header and data to Sheet1
sheet.cell(row=1, column=1, value='ATTACHMENT')
for index, value in enumerate(attachment_data, start=2):
    sheet.cell(row=index, column=1, value=value)

# Save the new Excel file
wb.save(excel_file_path)

print(f"Data from 'ATTACHMENT' column in CSV file has been added to 'Sheet1' in the new Excel file '{excel_file_name}'.")

# %%
print ("hello world")

# %%
#9th news download without chrome if internet there waiting
import http.client
import os
from urllib.parse import urlparse
import openpyxl
import threading
import time
import socket

# Path to the Excel file containing URLs
excel_file_path = 'Github/Excel format news/Symbols.xlsx'

# Function to check internet connection
def is_connected(host="8.8.8.8", port=53, timeout=3):
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error as ex:
        print(f"Internet connection lost: {ex}")
        return False

# Function to download a file with a 2-minute timeout
def download_file(url, save_path, timeout=120):
    try:
        # Ensure the directory exists or create it if it doesn't
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        # Parse the URL to get hostname and resource path
        parsed_url = urlparse(url)
        hostname = parsed_url.hostname
        resource_path = parsed_url.path

        # Connect to the server using HTTPS with a 2-minute timeout
        conn = http.client.HTTPSConnection(hostname, timeout=timeout)

        # Send a GET request to download the file
        conn.request("GET", resource_path)

        # Get the response from the server
        response = conn.getresponse()

        # Check if the request was successful (status code 200)
        if response.status == 200:
            # Open the file and write the response content (the file content)
            with open(save_path, 'wb') as f:
                f.write(response.read())
            print(f"Downloaded successfully to {save_path}")
        else:
            print(f"Failed to download file {os.path.basename(save_path)}. HTTP status code: {response.status}")

        # Close the connection
        conn.close()

    except Exception as e:
        # Handle any errors during the download process
        print(f"An error occurred while downloading {os.path.basename(save_path)}: {e}")

# Function to handle the downloading process for each URL
def process_url(row, url, save_path):
    print(f"Starting download for {url} at row {row}")
    
    # Check internet connection and wait if it's lost
    while not is_connected():
        print("Waiting for internet connection to be restored...")
        time.sleep(30)  # Wait for 30 seconds before checking again
    
    download_file(url, save_path)
    print(f"Completed download for {url} at row {row}")

# Load URLs from the Excel file and download each one using multi-threading
try:
    # Open the workbook in read-only mode
    wb = openpyxl.load_workbook(filename=excel_file_path, read_only=True)

    # Load the first sheet named 'Sheet1'
    sheet = wb['Sheet1']

    # Detect the last row with data in column 'A'
    max_row = sheet.max_row

    # List to keep track of threads
    threads = []

    # Iterate through the rows to get each URL and download the corresponding file
    for row in range(2, max_row + 1):  # Start from row 2 and go up to the last row with data
        cell_value = sheet[f'A{row}'].value  # Read value from column A
        if cell_value:  # Check if the cell is not empty
            url = cell_value.strip()  # Remove any extra spaces around the URL
            save_path = rf"C:\Users\bomma\Desktop\Github\All download pdfs\{os.path.basename(urlparse(url).path)}"
            
            # Create a new thread for each download
            thread = threading.Thread(target=process_url, args=(row, url, save_path))
            threads.append(thread)
            thread.start()  # Start the thread

            # Optional: Limit the number of concurrent threads to avoid overloading the system
            while threading.active_count() > 10:
                time.sleep(1)  # Wait for some threads to finish

    # Wait for all threads to complete
    for thread in threads:
        thread.join()

except Exception as e:
    # Handle errors that occur while reading the Excel file
    print(f"An error occurred while reading the Excel file: {e}")


# %%
#9 News extract 
#i update this 23/09/2025 bcz in excel file some time corrupt in linux so i remevode the part of code excel filedata read  see it is only data like this "News" like this data is there so mentioned this directly int the code 

import os
import pandas as pd
from PyPDF2 import PdfReader
import sqlite3
import time
import requests

# ------------------------
# Function: Extract text from PDF
# ------------------------
def extract_text_from_pdf(file_path, start_texts, end_texts, max_words=750):
    try:
        with open(file_path, 'rb') as file:
            reader = PdfReader(file)
            full_text = ""
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    full_text += text

            # Find start text
            start_index = -1
            for start_text in start_texts:
                start_index = full_text.find(start_text)
                if start_index != -1:
                    break

            # Find end text
            if start_index != -1:
                end_index = -1
                for end_text in end_texts:
                    end_index = full_text.find(end_text, start_index)
                    if end_index != -1:
                        break

                if end_index != -1:
                    extracted_text = full_text[start_index:end_index + len(end_text)]
                    words = extracted_text.split()
                    if len(words) > max_words:
                        extracted_text = ' '.join(words[:max_words])
                    return extracted_text

            return full_text
    except Exception as e:
        print(f"An error occurred while processing {file_path}: {e}")
        return None


# ------------------------
# Function: Update database
# ------------------------
def update_database_with_extracted_text(folder_path, db_file_path, table_name, start_texts, end_texts):
    try:
        conn = sqlite3.connect(db_file_path)
        c = conn.cursor()

        # Check if NEWS column exists, if not add it
        c.execute(f"PRAGMA table_info({table_name})")
        columns = c.fetchall()
        column_names = [col[1] for col in columns]
        if 'NEWS' not in column_names:
            c.execute(f"ALTER TABLE {table_name} ADD COLUMN NEWS TEXT")

        # Fetch rows
        c.execute(f"SELECT * FROM {table_name}")
        rows = c.fetchall()

        for row in rows:
            if len(row) > 10:  # make sure column index 10 exists
                pdf_file_name = str(row[10])
                pdf_file_path = os.path.join(folder_path, pdf_file_name)

                if os.path.exists(pdf_file_path):
                    extracted_text = extract_text_from_pdf(pdf_file_path, start_texts, end_texts)
                    if extracted_text is not None:
                        c.execute(f"UPDATE {table_name} SET NEWS = ? WHERE LINKS = ?",
                                  (extracted_text, pdf_file_name))
                        print(f"Text extracted successfully from: {pdf_file_name}")
                else:
                    print(f"PDF file '{pdf_file_name}' not found at path: {pdf_file_path}.")

        conn.commit()
        print("Update successful")
        conn.close()
    except Exception as e:
        print(f"An error occurred while updating the database: {e}")


# ------------------------
# Function: Send Telegram message
# ------------------------
def send_telegram_message(token, chat_id, message):
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": message
    }
    response = requests.post(url, json=payload)
    if response.status_code == 200:
        print("Message sent successfully!")
    else:
        print(f"Failed to send message. Status code: {response.status_code}")


# ------------------------
# File paths and parameters
# ------------------------
folder_path = "Github/All download pdfs"
db_file_path = "Github/Database/News Analysis.db"

# ✅ Directly mention table name (no Excel needed)
table_name = "News"

start_texts = ["Sub", "Subject"]
end_texts = [
    "Yours faithfully", "Yours sincerely", "Yours Faithfully",
    "Thanking you", "Thanking You", "Kindly take note of the same.",
    "Thanks & Regards"
]

# ------------------------
# Run the process
# ------------------------
update_database_with_extracted_text(folder_path, db_file_path, table_name, start_texts, end_texts)
time.sleep(10)
print("10 seconds have passed.")


# %%


# %%
#10 It is column name adding
import sqlite3

# Database file path
db_file_path = "Github/Database/News Analysis.db"

# Column names to be added
columns_to_add = [
    "News_date TEXT",
    "Date TEXT",
    "Market TEXT",
    "Market_date TEXT",
    "Previous_close TEXT",
    "Open TEXT",
    "High TEXT",
    "Low TEXT",
    "Close TEXT",
    "High_percentage TEXT",
    "Low_percentage TEXT",
    "Open_percentage TEXT",
    "Close_percentage TEXT",
    "d1_date TEXT",
    "d1_market_date TEXT",
    "d1_previous_close TEXT",
    "d1_open TEXT",
    "d1_high TEXT",
    "d1_low TEXT",
    "d1_close TEXT",
    "d1_high_percentage TEXT",
    "d1_low_percentage TEXT",
    "d1_open_percentage TEXT",
    "d1_close_percentage TEXT",
    "d7_date TEXT",
    "d7_market_date TEXT",
    "d7_previous_close TEXT",
    "d7_open TEXT",
    "d7_high TEXT",
    "d7_low TEXT",
    "d7_close TEXT",
    "d7_high_percentage TEXT",
    "d7_low_percentage TEXT",
    "d7_open_percentage TEXT",
    "d7_close_percentage TEXT",
    "Final_market_date TEXT",
    "Filter_dates TEXT",
    "Open_time TEXT",
    "Close_time TEXT",
    "d1_open_time TEXT",
    "d1_close_time TEXT",
    "d7_open_time TEXT",
    "d7_close_time TEXT",
    "Highest_time TEXT",
    "Lowest_time TEXT"
]

try:
    # Connect to the database
    conn = sqlite3.connect(db_file_path)
    cursor = conn.cursor()

    # Add new columns to the table
    for column in columns_to_add:
        cursor.execute(f"ALTER TABLE News ADD COLUMN {column};")

    # Commit the changes
    conn.commit()
    print("Columns added successfully.")

except sqlite3.Error as e:
    print("Error:", e)

finally:
    # Close the connection
    if conn:
        conn.close()



# %%
#11 Reciept Data copy
import sqlite3

# Database file path
db_file = "Github/Database/News Analysis.db"

try:
    # Connect to the database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Select data from the 'RECEIPT' column in the 'News' table
    cursor.execute("SELECT RECEIPT FROM News")

    # Fetch all rows
    rows = cursor.fetchall()

    # Iterate through the rows
    for row in rows:
        # Assuming the data in 'RECEIPT' is a string representing a date
        receipt_date = row[0]

        # Update the 'News_date' column with the value from 'RECEIPT'
        cursor.execute("UPDATE News SET News_date = ? WHERE RECEIPT = ?", (receipt_date, receipt_date))

    # Commit the changes
    conn.commit()
    print("Data updated successfully.")

except sqlite3.Error as e:
    print("Error:", e)

finally:
    # Close the connection
    conn.close()



# %%
#12 Date extract
import sqlite3
from datetime import datetime

# Path to your SQLite database file
db_file_path = "Github/Database/News Analysis.db"

# Function to extract date from datetime string
def extract_date(datetime_str):
    try:
        # Parse the datetime string
        datetime_obj = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M:%S')
        
        # Extract the date portion
        date_str = datetime_obj.strftime('%Y-%m-%d')
        
        return date_str
        
    except ValueError as e:
        print(f"Error extracting date from datetime string: {e}")
        return None

# Function to update the "Date" column in the "News" table
def update_date_column_in_news_table():
    try:
        # Connect to the database
        conn = sqlite3.connect(db_file_path)
        cursor = conn.cursor()
        
        # Check if the "News_date" column exists in the "News" table
        cursor.execute("PRAGMA table_info('News')")
        columns = cursor.fetchall()
        if any('News_date' in column for column in columns):
            # Fetch all rows from the "News" table
            cursor.execute("SELECT rowid, News_date FROM News")
            rows = cursor.fetchall()
            
            # Update each row with the extracted date
            for row in rows:
                row_id, news_date = row
                date = extract_date(news_date)
                if date:
                    # Update the "Date" column with the extracted date
                    cursor.execute("UPDATE News SET Date = ? WHERE rowid = ?", (date, row_id))
        else:
            print("The 'News' table does not have a 'News_date' column. Skipping...")
        
        # Commit the changes
        conn.commit()
        print("Date extracted and added to the 'Date' column in the 'News' table successfully.")
        
    except sqlite3.Error as e:
        print(f"Error updating 'Date' column in the 'News' table: {e}")
        
    finally:
        # Close the connection
        if conn:
            conn.close()

# Main function
def main():
    # Update the "Date" column in the "News" table
    update_date_column_in_news_table()

# Call the main function
if __name__ == "__main__":
    main()



# %%
#13 ofter and before market update
import sqlite3
from datetime import datetime

# Database file path
db_file = "Github/Database/News Analysis.db"

def update_market_column(db_file):
    # Connect to the database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    
    # Function to add Market column if not exists
    def add_market_column_if_not_exists(cursor):
        cursor.execute("PRAGMA table_info(News);")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]
        if 'Market' not in column_names:
            cursor.execute("ALTER TABLE News ADD COLUMN Market TEXT;")
            print("Added 'Market' column to table 'News'.")
    
    try:
        # Add Market column if not exists
        add_market_column_if_not_exists(cursor)
        
        # Retrieve data from the table
        cursor.execute("SELECT rowid, News_date FROM News;")
        rows = cursor.fetchall()
        
        # Update 'Market' column based on time conditions
        for row in rows:
            rowid = row[0]
            news_date_str = row[1]
            news_date = datetime.strptime(news_date_str, "%Y-%m-%d %H:%M:%S")
            news_time = news_date.time()
            market = ""
            if datetime.strptime("15:30:00", "%H:%M:%S").time() <= news_time <= datetime.strptime("23:59:00", "%H:%M:%S").time():
                market = "Ofter"
            elif datetime.strptime("09:15:00", "%H:%M:%S").time() <= news_time <= datetime.strptime("15:30:00", "%H:%M:%S").time():
                market = "Within"
            else:
                market = "Before"
            
            # Update 'Market' column for the row
            cursor.execute("UPDATE News SET Market = ? WHERE rowid = ?;", (market, rowid))
    except Exception as e:
        print(f"Error updating table 'News': {str(e)}")
    
    # Commit changes and close connection
    conn.commit()
    conn.close()

# Call the function to update the 'Market' column for the 'News' table
update_market_column(db_file)



# %%
#14 from database to csv empty files create
import sqlite3
import csv
import os

# Database file path
db_file_path = "Github/Database/News Analysis.db"

# Folder path to save CSV files
csv_folder_path = "Github/Mine Final CSVs"

# Connect to the database
conn = sqlite3.connect(db_file_path)
cursor = conn.cursor()

# Get symbols from the database
cursor.execute("SELECT DISTINCT SYMBOL FROM News")
symbols = cursor.fetchall()

# Create CSV files for each symbol
for symbol in symbols:
    symbol_name = symbol[0]
    csv_file_path = os.path.join(csv_folder_path, f"{symbol_name}.csv")
    
    # Create an empty CSV file
    with open(csv_file_path, "w", newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        # Write headers if needed
        # csv_writer.writerow(["Column1", "Column2", ...])
        
# Close the database connection
conn.close()

print("CSV files created successfully.")



# %% [markdown]
# #15 csv file data from database
# import sqlite3
# import csv
# import os
# 
# # Define paths
# database_path = r"C:\Users\bomma\Desktop\Rough Create\Single day Reports\Database\News Analysis.db"
# csv_folder_path = r"C:\Users\bomma\Desktop\Rough Create\Single day Reports\CSV file"
# 
# # Create CSV folder if it doesn't exist
# os.makedirs(csv_folder_path, exist_ok=True)
# 
# # Connect to the database
# conn = sqlite3.connect(database_path)
# cursor = conn.cursor()
# 
# # Select data from the "News" table
# cursor.execute("SELECT * FROM News")
# rows = cursor.fetchall()
# 
# # Define CSV file path
# csv_file_path = os.path.join(csv_folder_path, "news_data.csv")
# 
# # Write data to CSV file
# with open(csv_file_path, 'w', newline='') as csvfile:
#     csv_writer = csv.writer(csvfile)
#     # Write header
#     csv_writer.writerow([i[0] for i in cursor.description])
#     # Write rows
#     csv_writer.writerows(rows)
# 
# # Close database connection
# conn.close()
# 
# print("CSV file has been created successfully at:", csv_file_path)
# 
# 

# %%
#15 csv file data from database
import sqlite3
import csv
import os

# Define paths
database_path = "Github/Database/News Analysis.db"
csv_folder_path = "Github/CSV file"

# Create CSV folder if it doesn't exist
os.makedirs(csv_folder_path, exist_ok=True)

# Connect to the database
conn = sqlite3.connect(database_path)
cursor = conn.cursor()

# Select data from the "News" table
cursor.execute("SELECT * FROM News")

# Define CSV file path
csv_file_path = os.path.join(csv_folder_path, "news_data.csv")

# Write data to CSV file with UTF-8 encoding
with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
    csv_writer = csv.writer(csvfile)
    
    # Write header only once
    csv_writer.writerow([i[0] for i in cursor.description])
    
    # Write rows in chunks for memory efficiency
    while True:
        rows = cursor.fetchmany(1000)  # Fetch 1000 rows at a time
        if not rows:
            break
        csv_writer.writerows(rows)

# Close database connection
conn.close()

print("CSV file has been created successfully at:", csv_file_path)


# %%
#16 data add from Raw Csv file to each csv single file
import os
import pandas as pd

# File paths
source_csv_path = 'Github/CSV file/news_data.csv'
stocks_folder_path = 'Github/Mine Final CSVs'

# Read the source CSV file
source_df = pd.read_csv(source_csv_path)

# Get unique symbols from the 'SYMBOL' column
symbols = source_df['SYMBOL'].unique()

# Iterate over each symbol
for symbol in symbols:
    # Check if the corresponding CSV file exists for the symbol
    symbol_csv_path = os.path.join(stocks_folder_path, f'{symbol}.csv')
    if os.path.exists(symbol_csv_path):
        # Filter rows with the current symbol
        symbol_data = source_df[source_df['SYMBOL'] == symbol]
        
        # Append data to the symbol's CSV file
        if not os.path.getsize(symbol_csv_path):  # Check if the file is empty
            symbol_data.to_csv(symbol_csv_path, mode='a', header=True, index=False)  # Write header if file is empty
        else:
            symbol_data.to_csv(symbol_csv_path, mode='a', header=False, index=False)  # Append data without header
    else:
        print(f"CSV file not found for symbol '{symbol}'")



# %%
#17 empty database create
import sqlite3
import os

# Define the folder path
folder_path = "Github/News Now"

# Check if the folder exists, if not, create it
if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Construct the full file path for the SQLite database
db_file_path = os.path.join(folder_path, 'Final Analysis.db')

# Create an empty SQLite database file
with sqlite3.connect(db_file_path) as conn:
    print("Database file created successfully at:", db_file_path)

# Close the connection
conn.close()



# %% [markdown]
# #18 csv data save in database
# import os
# import pandas as pd
# import sqlite3
# 
# # Folder path containing CSV files
# csv_folder = r"C:\Users\bomma\Desktop\Github\Mine Final CSVs"
# 
# # Database file path
# db_file = r'C:\Users\bomma\Desktop\Github\News Now\Final Analysis.db'
# 
# # Function to create database table from CSV file
# def create_table_from_csv(csv_file, cursor):
#     table_name = os.path.splitext(os.path.basename(csv_file))[0]  # Extract table name from CSV file name
#     df = pd.read_csv(csv_file)  # Read CSV file into DataFrame
#     try:
#         # Replace any spaces in column names with underscores
#         df.columns = df.columns.str.replace(' ', '_')
#     except ValueError:
#         pass  # Ignore length mismatch error
#     df.to_sql(table_name, conn, if_exists='replace', index=False)  # Write DataFrame to SQLite database as a table
# 
# # Connect to SQLite database
# conn = sqlite3.connect(db_file)
# cursor = conn.cursor()
# 
# # Iterate over CSV files in the folder
# for file_name in os.listdir(csv_folder):
#     if file_name.endswith('.csv'):
#         csv_file_path = os.path.join(csv_folder, file_name)
#         create_table_from_csv(csv_file_path, cursor)
# 
# # Commit changes and close connection
# conn.commit()
# conn.close()
# 
# print("Database successfully created from CSV files.")
# 
# 

# %%
#Updated version 18 csv data save in database  (Python int too large to convert to SQLite INTEGER some tables notr taking data it shows error so ) 
import os
import pandas as pd
import sqlite3
import traceback

# Folder path containing CSV files
csv_folder = "Github/Mine Final CSVs"

# Database file path
db_file = 'Github/News Now/Final Analysis.db'

# Function to create database table from CSV file
def create_table_from_csv(csv_file, conn):
    table_name = os.path.splitext(os.path.basename(csv_file))[0]  # Extract table name from CSV file name
    try:
        df = pd.read_csv(csv_file)  # Read CSV file into DataFrame
        
        # Replace spaces in column names with underscores
        df.columns = df.columns.str.replace(' ', '_')
        
        # Check for large integers and convert them to strings if needed
        for column in df.select_dtypes(include=['int']).columns:
            if df[column].max() > 9223372036854775807:  # SQLite INTEGER max value
                df[column] = df[column].astype(str)  # Convert to string
        
        # Write DataFrame to SQLite database as a table
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        print(f"Table '{table_name}' created successfully.")
    except Exception as e:
        # Print the error and the file causing it
        print(f"Error processing file: {csv_file}")
        print("Error details:")
        traceback.print_exc()

# Connect to SQLite database
conn = sqlite3.connect(db_file)

# Iterate over CSV files in the folder
for file_name in os.listdir(csv_folder):
    if file_name.endswith('.csv'):
        csv_file_path = os.path.join(csv_folder, file_name)
        print(f"Processing file: {csv_file_path}")
        create_table_from_csv(csv_file_path, conn)

# Commit changes and close connection
conn.commit()
conn.close()

print("Database creation process completed.")


# %%
print("From Here Main Analysis Start")


# %%
import zipfile
import os

# --- Source file path ---
db_path = "Github/News Now/Final Analysis.db"

# --- Destination folder path ---
zip_folder = "Github/Final Zip"

# --- Ensure the folder exists ---
os.makedirs(zip_folder, exist_ok=True)

# --- Create ZIP file path in destination folder ---
zip_name = "Final Analysis.zip"
zip_path = os.path.join(zip_folder, zip_name)

# --- Compress the file ---
with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
    zipf.write(db_path, os.path.basename(db_path))  # add only filename, not full path

print(f"File zipped successfully: {zip_path}")


# %%
# Telegram message + file sender for GitHub Actions (reads secrets from env)
import os
import requests

# Read from environment variables set by GitHub Actions
BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")
FILE_PATH = os.getenv("FILE_PATH", "Final Zip/Final Analysis.zip")  # default relative path

MESSAGE = (
    "GitHub Rashmi: News Download and Extraction completed.\n"
    "SQLite database file attached.\n\n"
    "Thank you sir,\nSincerely,\nJanavi K"
)

def send_telegram_message(token, chat_id, message):
    """Send a Telegram text message."""
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {"chat_id": chat_id, "text": message}
    resp = requests.post(url, json=payload, timeout=15)
    resp.raise_for_status()
    print("Message sent successfully")

def send_telegram_file(token, chat_id, file_path, caption=None):
    """Send a file to Telegram chat."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as f:
        files = {"document": f}
        data = {"chat_id": chat_id, "caption": caption or ""}
        resp = requests.post(url, data=data, files=files, timeout=120)
        resp.raise_for_status()
    print("File sent successfully")

if __name__ == "__main__":
    if not BOT_TOKEN or not CHAT_ID:
        raise SystemExit("Missing TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID env vars")
    send_telegram_message(BOT_TOKEN, CHAT_ID, MESSAGE)
    send_telegram_file(BOT_TOKEN, CHAT_ID, FILE_PATH, caption="Final Analysis Database File")


# %%
print('from here only for laptop code run this is copy the files to google drive for access mobile codes')

# %%



